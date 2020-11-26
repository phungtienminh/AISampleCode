from openpyxl import workbook, load_workbook
from openpyxl.styles import Alignment

import pandas as pd
import numpy as np

from sklearn.linear_model import LinearRegression

# define a constant epsilon
eps = 1e-6

wb = load_workbook("./3.TBA_Round02-Working Spreadsheet.xlsx")
sheets = wb.sheetnames # ['Productivity-Benchmark', 'Assumption-LIC', 'Assumption-TGB', 'Summary', 'IRR', 'Glossary']

# ================================================================================
# Task A solver

#print(sheets)
assumption_tgb_sheet = wb[sheets[2]]
#print(assumption_tgb_sheet.cell(row = 1, column = 2).value)

'''
Retrieve all important data in assumption-TGB sheet
'''

total_sale_staff = []
monthly_active_ratio = []
case_per_active_per_month = []
average_case_size = []
total_bank_customer = []

product_mix = np.array([0.1, 0.25, 0.5, 0.15])
nb_profit_margin = np.array([0.8, 0.3, 0.3, 0.6])
avg_nb_profit_margin = round(np.dot(product_mix, nb_profit_margin), 2)

for r in range(3, 7):
    for c in range(2, 17):
        if r == 3:
            total_sale_staff.append(assumption_tgb_sheet.cell(row = r, column = c).value)
        elif r == 4:
            monthly_active_ratio.append(assumption_tgb_sheet.cell(row = r, column = c).value)
        elif r == 5:
            case_per_active_per_month.append(assumption_tgb_sheet.cell(row = r, column = c).value)
        else:
            average_case_size.append(assumption_tgb_sheet.cell(row = r, column = c).value)

for c in range(2, 17):
    total_bank_customer.append(assumption_tgb_sheet.cell(row = 18, column = c).value)

#print(total_sale_staff)
#print(monthly_active_ratio)
#print(case_per_active_per_month)
#print(average_case_size)
total_sale_staff = pd.Series(total_sale_staff, index = range(2021, 2036))
monthly_active_ratio = pd.Series(monthly_active_ratio, index = range(2021, 2036))
case_per_active_per_month = pd.Series(case_per_active_per_month, index = range(2021, 2036))
average_case_size = pd.Series(average_case_size, index = range(2021, 2036))
total_bank_customer = pd.Series(total_bank_customer, index = range(2021, 2036))

# Cleaning data
average_case_size = average_case_size.apply(lambda x: round(x, 0)) # for consistency with data in spreadsheet
total_bank_customer = total_bank_customer.apply(lambda x: round(x, 0)) # remove scientific notation
#print(total_bank_customer)

# Assumption-LIC sheet
assumption_lic_sheet = wb[sheets[1]]
fixed_cost = []
variable_cost = []
risk_discount_rate = assumption_lic_sheet.cell(row = 6, column = 2).value
annual_policy_lapse_rate = assumption_lic_sheet.cell(row = 7, column = 2).value

for r in range(3, 5):
    for c in range(2, 17):
        if r == 3:
            fixed_cost.append(assumption_lic_sheet.cell(row = r, column = c).value)
        else:
            variable_cost.append(assumption_lic_sheet.cell(row = r, column = c).value)

fixed_cost = pd.Series(fixed_cost, index = range(2021, 2036)).apply(lambda x: -x)
variable_cost = pd.Series(variable_cost, index = range(2021, 2036)).apply(lambda x: round(x + eps, 3)) # avoid rounding error

# Fill the answer
summary_sheet = wb[sheets[3]]

def fill_data_A():
    policy_sold_inforce = 0
    deal_value = 0

    for c in range(4, 19):
        policy_sold = total_sale_staff.loc[c + 2017] * monthly_active_ratio.loc[c + 2017] * case_per_active_per_month.loc[c + 2017] * 12
        policy_sold = round(policy_sold, 0)
        total_anp = policy_sold * average_case_size.loc[c + 2017]
        gross_profit = total_anp * avg_nb_profit_margin
        additional_cost = fixed_cost.loc[c + 2017] + total_anp * variable_cost.loc[c + 2017]
        net_profit = gross_profit - additional_cost
        deal_value += net_profit

        if c == 4:
            policy_sold_inforce = policy_sold
        else:
            policy_sold_inforce = policy_sold_inforce * (1.0 - annual_policy_lapse_rate) + policy_sold

        customer_penetration = policy_sold_inforce / total_bank_customer.loc[c + 2017]

        summary_sheet.cell(row = 3, column = c).value = policy_sold
        summary_sheet.cell(row = 4, column = c).value = total_anp / 1000
        summary_sheet.cell(row = 5, column = c).value = gross_profit / 1000
        summary_sheet.cell(row = 6, column = c).value = additional_cost / 1000
        summary_sheet.cell(row = 7, column = c).value = net_profit / 1000
        summary_sheet.cell(row = 10, column = c).value = policy_sold_inforce
        summary_sheet.cell(row = 11, column = c).value = str(int(round(customer_penetration * 100, 0))) + "%"
        summary_sheet.cell(row = 11, column = c).alignment = Alignment(horizontal = "right")

    deal_value = deal_value * (1.0 - risk_discount_rate)
    summary_sheet.cell(row = 13, column = 4).value = deal_value / 1000

fill_data_A()
# ================================================================================

# ================================================================================
# Task B solver
productivity_bmb_sheet = wb[sheets[0]]

# Should not use 2D array here to facilitate dataframe usage
q1_2018, q2_2018, q3_2018, q4_2018 = [], [], [], []
q1_2019, q2_2019, q3_2019, q4_2019 = [], [], [], []
q1_2020, q2_2020, q3_2020, q4_2020 = [], [], [], []

for c in range(2, 14):
    for r in range(9, 509):
        val = productivity_bmb_sheet.cell(row = r, column = c).value
        if c == 2:
            q1_2018.append(val)
        elif c == 3:
            q2_2018.append(val)
        elif c == 4:
            q3_2018.append(val)
        elif c == 5:
            q4_2018.append(val)
        elif c == 6:
            q1_2019.append(val)
        elif c == 7:
            q2_2019.append(val)
        elif c == 8:
            q3_2019.append(val)
        elif c == 9:
            q4_2019.append(val)
        elif c == 10:
            q1_2020.append(val)
        elif c == 11:
            q2_2020.append(val)
        elif c == 12:
            q3_2020.append(val)
        else:
            q4_2020.append(val)

policy_sold_df = pd.DataFrame({"Q12018": q1_2018, "Q22018": q2_2018, "Q32018": q3_2018, "Q42018": q4_2018, \
                               "Q12019": q1_2019, "Q22019": q2_2019, "Q32019": q3_2019, "Q42019": q4_2019, \
                               "Q12020": q1_2020, "Q22020": q2_2020, "Q32020": q3_2020, "Q42020": q4_2020}, \
                               index = range(1, 501))

q1_2018, q2_2018, q3_2018, q4_2018 = [], [], [], []
q1_2019, q2_2019, q3_2019, q4_2019 = [], [], [], []
q1_2020, q2_2020, q3_2020, q4_2020 = [], [], [], []

for c in range(17, 29):
    for r in range(9, 509):
        val = productivity_bmb_sheet.cell(row = r, column = c).value
        if c == 17:
            q1_2018.append(val)
        elif c == 18:
            q2_2018.append(val)
        elif c == 19:
            q3_2018.append(val)
        elif c == 20:
            q4_2018.append(val)
        elif c == 21:
            q1_2019.append(val)
        elif c == 22:
            q2_2019.append(val)
        elif c == 23:
            q3_2019.append(val)
        elif c == 24:
            q4_2019.append(val)
        elif c == 25:
            q1_2020.append(val)
        elif c == 26:
            q2_2020.append(val)
        elif c == 27:
            q3_2020.append(val)
        else:
            q4_2020.append(val)


case_size_df = pd.DataFrame({"Q12018": q1_2018, "Q22018": q2_2018, "Q32018": q3_2018, "Q42018": q4_2018, \
                             "Q12019": q1_2019, "Q22019": q2_2019, "Q32019": q3_2019, "Q42019": q4_2019, \
                             "Q12020": q1_2020, "Q22020": q2_2020, "Q32020": q3_2020, "Q42020": q4_2020}, \
                             index = range(1, 501))

total_sale_staff_B = total_sale_staff.copy()
monthly_active_ratio_B = monthly_active_ratio.copy()
case_per_active_per_month_B = case_per_active_per_month.copy()
average_case_size_B = average_case_size.copy()

# monthly_active_ratio_B
for y in range(2018, 2021):
    avg_active_ratio = 0
    for q in range(1, 5):
        col_name = "Q{}{}".format(q, y)
        cnt_active = 0
        for r in range(1, 501):
            if policy_sold_df.loc[r, col_name] > 0:
                cnt_active += 1

        avg_active_ratio += cnt_active / 500

    avg_active_ratio /= 4
    monthly_active_ratio_B.loc[y + 3] = avg_active_ratio

# case_per_active_per_month_B
for y in range(2018, 2021):
    avg_case_per_month = 0
    for q in range(1, 5):
        col_name = "Q{}{}".format(q, y)
        cnt_case = 0
        cnt_active = 0
        for r in range(1, 501):
            if policy_sold_df.loc[r, col_name] > 0:
                cnt_case += policy_sold_df.loc[r, col_name] // 3
                cnt_active += 1

        avg_case_per_month += cnt_case / cnt_active

    avg_case_per_month /= 4
    case_per_active_per_month_B.loc[y + 3] = avg_case_per_month

# average_case_size_B
for y in range(2018, 2021):
    avg_case_size = 0
    for q in range(1, 5):
        col_name = "Q{}{}".format(q, y)
        cnt_case_size = 0
        cnt_active = 0
        for r in range(1, 501):
            if case_size_df.loc[r, col_name] > 0:
                cnt_case_size += case_size_df.loc[r, col_name]
                cnt_active += 1

        avg_case_size += cnt_case_size / cnt_active

    avg_case_size /= 4
    average_case_size_B.loc[y + 3] = avg_case_size

# Change value according to problem statement
monthly_active_ratio_B.loc[2030] = 0.5
monthly_active_ratio_B.loc[2035] = 0.6
case_per_active_per_month_B.loc[2030] = 6.0
case_per_active_per_month_B.loc[2035] = 7.0
average_case_size_B.loc[2030] = 1200
average_case_size_B.loc[2035] = 1500

monthly_active_ratio_data = np.array([1, 2, 3, 10, 15]).reshape((-1, 1))
monthly_active_ratio_label = np.array([monthly_active_ratio_B.loc[2021], monthly_active_ratio_B.loc[2022], \
                                      monthly_active_ratio_B.loc[2023], monthly_active_ratio_B.loc[2030], \
                                      monthly_active_ratio_B.loc[2035]])

case_per_active_per_month_data = np.array([1, 2, 3, 10, 15]).reshape((-1, 1))
case_per_active_per_month_label = np.array([case_per_active_per_month_B.loc[2021], \
                                            case_per_active_per_month_B.loc[2022], \
                                            case_per_active_per_month_B.loc[2023], \
                                            case_per_active_per_month_B.loc[2030], \
                                            case_per_active_per_month_B.loc[2035]])

average_case_size_data = np.array([1, 2, 3, 10, 15]).reshape((-1, 1))
average_case_size_label = np.array([average_case_size_B.loc[2021], average_case_size_B.loc[2022], \
                                    average_case_size_B.loc[2023], average_case_size_B.loc[2030], \
                                    average_case_size_B.loc[2035]])

# monthly_active_ratio model

model_monthly_active_ratio = LinearRegression()
model_monthly_active_ratio.fit(monthly_active_ratio_data, monthly_active_ratio_label)
#print("Score:", model_monthly_active_ratio.score(monthly_active_ratio_data, monthly_active_ratio_label))
#print("Intercept:", model_monthly_active_ratio.intercept_)
#print("Slope:", model_monthly_active_ratio.coef_[0])

monthly_active_ratio_test = np.array([4, 5, 6, 7, 8, 9, 11, 12, 13, 14])
monthly_active_ratio_pred = model_monthly_active_ratio.predict(monthly_active_ratio_test.reshape((-1, 1)))

for i in range(len(monthly_active_ratio_test)):
    key = monthly_active_ratio_test[i] + 2020
    val = monthly_active_ratio_pred[i]
    monthly_active_ratio_B.loc[key] = val

# case_per_active_per_month model
model_case_per_active_per_month = LinearRegression()
model_case_per_active_per_month.fit(case_per_active_per_month_data, case_per_active_per_month_label)
case_per_active_per_month_test = np.array([4, 5, 6, 7, 8, 9, 11, 12, 13, 14])
case_per_active_per_month_pred = model_case_per_active_per_month.predict(case_per_active_per_month_test.reshape((-1, 1)))

for i in range(len(case_per_active_per_month_test)):
    key = case_per_active_per_month_test[i] + 2020
    val = case_per_active_per_month_pred[i]
    case_per_active_per_month_B.loc[key] = val

# average_case_size model
model_average_case_size = LinearRegression()
model_average_case_size.fit(average_case_size_data, average_case_size_label)
average_case_size_test = np.array([4, 5, 6, 7, 8, 9, 11, 12, 13, 14])
average_case_size_pred = model_average_case_size.predict(average_case_size_test.reshape((-1, 1)))

for i in range(len(average_case_size_test)):
    key = average_case_size_test[i] + 2020
    val = average_case_size_pred[i]
    average_case_size_B.loc[key] = val

def fill_data_B():
    policy_sold_inforce = 0
    deal_value = 0

    for c in range(4, 19):
        policy_sold = total_sale_staff_B.loc[c + 2017] * monthly_active_ratio_B.loc[c + 2017] * case_per_active_per_month_B.loc[c + 2017] * 12
        policy_sold = round(policy_sold, 0)
        total_anp = policy_sold * average_case_size_B.loc[c + 2017]
        gross_profit = total_anp * avg_nb_profit_margin
        additional_cost = fixed_cost.loc[c + 2017] + total_anp * variable_cost.loc[c + 2017]
        net_profit = gross_profit - additional_cost
        deal_value += net_profit

        if c == 4:
            policy_sold_inforce = policy_sold
        else:
            policy_sold_inforce = policy_sold_inforce * (1.0 - annual_policy_lapse_rate) + policy_sold

        customer_penetration = policy_sold_inforce / total_bank_customer.loc[c + 2017]

    deal_value = deal_value * (1.0 - risk_discount_rate)
    summary_sheet.cell(row = 20, column = 4).value = deal_value / 1000

fill_data_B()
# ================================================================================

# ================================================================================
# Task C solver
irr_sheet = wb[sheets[4]]

# Change value according to problem statement
product_mix_C = np.array([0.1, 0.45, 0.4, 0.05])
nb_profit_margin_C = np.array([0.75, 0.25, 0.25, 0.55])
avg_nb_profit_margin_C = round(np.dot(product_mix_C, nb_profit_margin_C), 2)
fixed_cost_C = fixed_cost.copy()
variable_cost_C = variable_cost.copy()

for y in range(2021, 2036):
    if y <= 2022:
        fixed_cost_C.loc[y] += 2000000
    else:
        fixed_cost_C.loc[y] += 500000

variable_cost_C = variable_cost_C.apply(lambda x: x - 0.025)

def standardize_str(s):
    t = ""
    cnt = 0
    for ch in s[::-1]:
        t += ch
        cnt += 1

        if cnt % 3 == 0 and ch != '-':
            cnt = 0
            t += ','

    return t[::-1]

def fill_data_C():
    policy_sold_inforce = 0
    deal_value = 0
    total_anp_array = []

    for c in range(4, 19):
        policy_sold = total_sale_staff_B.loc[c + 2017] * monthly_active_ratio_B.loc[c + 2017] * case_per_active_per_month_B.loc[c + 2017] * 12
        policy_sold = round(policy_sold, 0)
        total_anp = policy_sold * average_case_size_B.loc[c + 2017] * 0.8
        gross_profit = total_anp * avg_nb_profit_margin_C
        additional_cost = fixed_cost_C.loc[c + 2017] + total_anp * variable_cost_C.loc[c + 2017]
        net_profit = gross_profit - additional_cost
        deal_value += net_profit

        total_anp_array.append(total_anp)

        if c == 4:
            policy_sold_inforce = policy_sold
        else:
            policy_sold_inforce = policy_sold_inforce * (1.0 - annual_policy_lapse_rate) + policy_sold

        customer_penetration = policy_sold_inforce / total_bank_customer.loc[c + 2017]

        summary_sheet.cell(row = 26, column = c).value = total_anp / 1000
        summary_sheet.cell(row = 27, column = c).value = gross_profit / 1000
        summary_sheet.cell(row = 28, column = c).value = additional_cost / 1000

        if net_profit < 0:
            summary_sheet.cell(row = 29, column = c).value = standardize_str("-" + str(int(round(-net_profit / 1000, 0))))
            summary_sheet.cell(row = 29, column = c).alignment = Alignment(horizontal = "right")
        else:
            summary_sheet.cell(row = 29, column = c).value = net_profit / 1000

    deal_value = deal_value * (1.0 - risk_discount_rate)
    summary_sheet.cell(row = 32, column = 4).value = deal_value / 1000
    summary_sheet.cell(row = 33, column = 4).value = 38804

    for r in range(3, 18):
        irr_sheet.cell(row = r, column = 3).value = total_anp_array[r - 3] / 1000

    irr_sheet.cell(row = 3, column = 6).value = 38804

fill_data_C()
# ================================================================================
# Save my work
wb.save("./answer.xlsx")
